# -*- coding: utf-8 -*-
def excel_to_frame(path,sheets=[],header=1,index_col=2,pruduct_name_col=0,data_breaks=1,header_info=0,index_col_info=0,sheet_info=u'产品信息'):
    import pandas as pd
    import numpy as np
    if not sheets:
        import openpyxl
        wb=openpyxl.load_workbook(path)
        sheets=wb.sheetnames
        sheets.remove(sheet_info)
    product_data={}
    for x in sheets:
        sheet_data=pd.read_excel(path,sheet_name=x,header=header,index_col=index_col)
        data_rows=len(sheet_data.index.tolist())
        data_cols=1
        while sheet_data.columns[index_col+data_cols+data_breaks]!=(sheet_data.columns[index_col]+'.1'):
            data_cols +=1
        producct_columns=pd.read_excel(path,sheet_name=x,header=header-1).iloc[0,index_col+1:index_col+data_cols].tolist()
        product_names=pd.read_excel(path,sheet_name=x,header=None).iloc[:,pruduct_name_col].dropna()
        product_id=0
        for x in product_names:
            product_DataFrame=sheet_data.iloc[0:data_rows,(data_cols+data_breaks)*product_id+index_col:(data_cols+data_breaks)*product_id+index_col+data_cols-1]
            product_DataFrame.columns=producct_columns
            product_id+=1
            product_data[x] = product_DataFrame
    product_info=pd.read_excel(path,sheet_name=sheet_info,header=header_info,index_col=index_col_info)
    return product_data,product_info

def excel1_to_frame(path,sheet_info=u'产品信息',sheet_data=u'数据',index_name=u'日期',data_name=u'产品'):
    import pandas as pd
    info=pd.read_excel(path,sheet_name=sheet_info,index_col=0)
    info_name=info[data_name]
    data_sheet=pd.read_excel(path,sheet_name=sheet_data,header=0,index_col=[0,1])
    data_name=data_sheet.unstack().index.tolist()
    data={}
    for x in info_name:
        frame=data_sheet.loc[x].unstack().unstack()
        frame.index.names=[index_name]
        frame.columns.names=[None]
        data[x]=frame
    return data,info

def frame_to_excel1(data,info,path,sheet_info=u'产品信息',sheet_data=u'数据',index_name=u'指标',data_name=u'产品'):
    import pandas as pd
    panel=pd.Panel(data)
    data_multi=panel.to_frame(filter_observations=False)
    col_name=data_multi.index.names[0]
    data_multi.index.names=[col_name,index_name]
    data_multi.columns.names=[data_name]
    data_multi=data_multi.unstack(col_name).stack(data_name,dropna=False).swaplevel(data_name,index_name).sort_index()
    data_multi.columns.names=[None]
    data_writer = pd.ExcelWriter(path)
    info.to_excel(data_writer,sheet_info)
    data_multi.to_excel(data_writer,sheet_data)
    data_writer.save()
    return None

def sql_to_frame(login,sheet_info=u'产品信息',index_name=u'日期'):
    import pandas as pd
    import numpy as np
    import re
    from sqlalchemy import create_engine
    import mysql.connector
    re_login=re.match(r'.*//(.*):(.*)@(.*)/(\w*)',login)
    user=re_login.group(1)
    password=re_login.group(2)
    host=re_login.group(3)
    database=re_login.group(4)
    con_imfo = mysql.connector.connect(user=user, password=password, host=host, database='information_schema')
    cursor_imfo = con_imfo.cursor()
    cursor_imfo.execute("select TABLE_NAME from tables where TABLE_SCHEMA='%s'" %database)
    product_names = [x[0] for x in cursor_imfo.fetchall()]#每一条结果都是一个tuple,因此需要转化为list.
    product_names.remove(sheet_info)
    con = create_engine(login)
    product_data={}
    for x in product_names:
        product_DataFrame=pd.read_sql_table(x, con, index_col=index_name)
        product_data[x] = product_DataFrame
    product_info=pd.read_sql_table(sheet_info, con, index_col='ID')
    return product_data,product_info

def frame_to_sql(data,info,login,sheet_info=u'产品信息'):
    from sqlalchemy import create_engine
    con = create_engine(login)
    for x in data:
        data[x].to_sql(x,con,if_exists='replace')
    info.to_sql(sheet_info,con,if_exists='replace')
    return None

def sql_to_frame1(login,sheet_info=u'产品信息'):
    import pandas as pd
    import numpy as np
    import re
    from sqlalchemy import create_engine
    import mysql.connector
    re_login=re.match(r'.*//(.*):(.*)@(.*)/(\w*)',login)
    user=re_login.group(1)
    password=re_login.group(2)
    host=re_login.group(3)
    database=re_login.group(4)
    con_imfo = mysql.connector.connect(user=user, password=password, host=host, database='information_schema')
    cursor_imfo = con_imfo.cursor()
    cursor_imfo.execute("select TABLE_NAME from tables where TABLE_SCHEMA='%s'" %database)
    product_names = [x[0] for x in cursor_imfo.fetchall()]#每一条结果都是一个tuple,因此需要转化为list.
    product_names.remove(sheet_info)
    con = create_engine(login)
    product_data={}
    for x in product_names:
        product_DataFrame=pd.read_sql_table(x, con)
        product_data[x] = product_DataFrame
    product_info=pd.read_sql_table(sheet_info, con, index_col='ID')
    return product_data,product_info

def frame1_to_excel(data,info,path,header=1,index_col=2,pruduct_name_col=0,sheet_info=u'产品信息',kind_name=u'合作方式',data_name=u'产品'):
    import pandas as pd
    import numpy as np
    import openpyxl
    #获取合作方式与产品对应的的Series
    info_kinds=info.set_index(kind_name).loc[:,data_name]
    #获取合作方式的种类
    info_kind=set(info_kinds.keys())
    #对每个DataFrame进行修改以便于放入excel
    for x in data:
        data[x]['None']=''
        data[x]['id']=range(len(data[x]))
        data[x]=data[x].set_index('id')
    #将不同类别的DataFrame连接在一起
    product_kind_frame={}
    for x in info_kind:
        product_kind_frame_separate=[]
        for y in info_kinds[x]:
            product_kind_frame_separate.append(data[y])
        product_kind_frame[x]=pd.concat(product_kind_frame_separate, axis = 1)
    #将三个类别的表放入对应的sheet中
    data_writer = pd.ExcelWriter(path)
    info.to_excel(data_writer,sheet_info)
    for x in product_kind_frame:
        product_kind_frame[x].to_excel(data_writer,x,startrow=header ,startcol=index_col-1)
    data_writer.save()
    #稍微再进行修改
    #注意,excel的行数列数都是从1开始数的
    wb=openpyxl.load_workbook(path)
    for x in info_kind:
        ws=wb[x]
        product_kind_names=list(info_kinds[x])
        columns_each_product_diff=len(data[product_kind_names[0]].columns)
        rows_each_product=len(data[product_kind_names[0]].index)
        for i in range(1,len(product_kind_names)+1):
            ws.cell(row=i , column=pruduct_name_col+1).value=product_kind_names[i-1]
            ws.cell(row=header , column=(i-1)*columns_each_product_diff+index_col+1).value=product_kind_names[i-1]
            ws.cell(row=header+1 , column=i*columns_each_product_diff+index_col).value=''
        for i in range(1,rows_each_product+header+2):
            ws.cell(row=i, column=index_col).value=''
    wb.save(path)    
    return None

def sql_to_excel(login,path,header=1,index_col=2,pruduct_name_col=0,sheet_info=u'产品信息',kind_name=u'合作方式',data_name=u'产品'):
    data,info=sql_to_frame1(login=login,sheet_info=sheet_info)
    frame1_to_excel(data=data,info=info,path=path,header=header,index_col=index_col,pruduct_name_col=pruduct_name_col,sheet_info=sheet_info,kind_name=kind_name,data_name=data_name)
    return None
